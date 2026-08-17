#定稿版本

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from xzdict1 import XZ1
from xzdict2 import XZ2
from openpyxl.styles import PatternFill

print("123")
df = pd.read_excel("点检.xlsx")

# 语法:df["列名"].unique()
# 作用:获取"产线"这一列中所有不重复的值,返回一个数组,用于后续遍历每条产线
all_lines = df["产线"].unique()

df["创建时间"] = pd.to_datetime(df["创建时间"])
df["实际点检时间"] = pd.to_datetime(df["实际点检时间"])
df["计划点检时间"] = pd.to_datetime(df["计划点检时间"])

A_START = pd.to_datetime("2026-06-20 00:00:00")

A_END = pd.to_datetime("2026-06-20 09:30:00")

B_START = pd.to_datetime("2026-06-20 10:00:00")

B_END = pd.to_datetime("2026-06-20 23:59:59")

condition_a = (df["创建时间"] >= A_START) & (df["创建时间"] <= A_END)

condition_b = (df["创建时间"] >= B_START) & (df["创建时间"] <= B_END)

# 语法:df.loc[行条件, 列名] = 值
# 作用:满足A班条件的行,在"班次"列赋值为"A班"
df.loc[condition_a, "班次"] = "白班"
# 满足B班条件的行,在"班次"列赋值为"B班"
df.loc[condition_b, "班次"] = "夜班"

# 定义班次列表,用于后面循环遍历A班、B班
BANCI = ["白班", "夜班"]
# 定义点检周期:日、周、月
cycle_types = [["日", "日检"], ["周", "周检"], ["月", "月检"]]
# 定义空列表,用于存储所有统计结果
result = []

file2 = "账号表格.xlsx"

col_people = "*点检人员"

col_name = "用户姓名"
col_account = "登录账号"

df222 = pd.read_excel(file2)

# 姓名 → 账号 映射
name_to_account = dict(zip(df222[col_name], df222[col_account]))

# ===================== 3. 循环统计核心逻辑 =====================

for line in all_lines:
    # 遍历每个班次(A班、B班)
    for banci in BANCI:
        # 每次循环都创建空字典,存储当前产线+班次的统计数据
        # 作用:防止上一次循环的数据残留
        stats = {}

        # 语法:df[(条件1) & (条件2)]
        # 作用:筛选出 当前产线 且 当前班次 的所有数据,赋值给file
        file = df[(df["产线"] == line) & (df["班次"] == banci)]

        # 遍历日、周、月,分别统计
        for cycle in cycle_types:
            col = file["点检计划"].fillna("").astype(str).str.split("/").str[-1]
            shuju = file[col.isin(cycle)]

            # 语法:len(数据)
            # 作用:获取数据行数 = 计划点检数量
            plan_count = len(shuju)
            # print(plan_count)
            # 语法:列.notna().sum()
            # 作用:统计"实际点检时间"不为空的数量 = 实际完成数量
            # actual_count = shuju["实际点检时间"].notna().sum()
            actual_count = (shuju["点检进度"].astype(str) == "已完成").sum()
            # print(actual_count)
            # 语法:字典[键名] = 值
            # 作用:把统计结果存入字典
            stats[f"计划{cycle[0]}点检量"] = plan_count
            stats[f"实际{cycle[0]}点检量"] = actual_count
            stats[f"差异{cycle[0]}点检量"] = plan_count - actual_count

        # 语法:列.isna()
        # 作用:筛选出"实际点检时间"为空的数据(未完成点检)
        # incomplete = file[file["实际点检时间"].isna()]
        incomplete = file[file["点检进度"].astype(str).isin(["待点检", "未开始"])]


        # =================================================================

        # 转换函数（支持 ; | 、 多种分隔符）
        def replace_names(text):
            if pd.isna(text):
                return ""
            text = str(text).strip()

            # 统一分隔符
            for sep in ["#", ";", "、"]:
                text = text.replace(sep, ",")
            parts = text.split(",")

            # 替换账号
            zhanghao = []
            for name in parts:
                name = name.strip()
                acc = name_to_account.get(name, name)  # 找不到就返回原名
                if acc:
                    zhanghao.append(acc)

            return "、".join(zhanghao)


        acc_series = incomplete[col_people].apply(replace_names)
        all_acc_texts = acc_series.dropna().unique()
        account_str = ' 、'.join(all_acc_texts) if len(all_acc_texts) > 0 else '-'

        # =================================================================

        # 语法:数据.get(列名, 备用值).dropna().unique()
        # 作用:获取未完成人员列,去掉空值,去重
        personnel = incomplete.get("*点检人员", pd.Series()).dropna().unique()

        # 语法:条件判断式
        # 作用:有人员则用分号连接,无人员则显示"-"
        personnel_name = '  #'.join(personnel) if len(personnel) > 0 else '-'
        # print(personnel_name)

        # 语法:列表.append(字典)
        # 作用:把当前产线+班次的所有统计结果,添加到总结果列表中
        result.append({
            "产线": line,
            "班次": banci,
            **stats,  # 把stats字典里的所有键值对展开进来
            "未完成设备点检姓名": personnel_name,
            "未完成设备点检账号": account_str,
            "未完成设备点检原因": "-",
            "线长": XZ1[line][banci],
            "模块线长": "-"
        })

# ===================== 4. 生成统计表格 + 增加模块列 =====================
# 语法:pd.DataFrame(列表)
# 作用:把result列表转换为DataFrame表格
dn = pd.DataFrame(result)

# 语法:df["新列名"] = df["列名"].str[:2]
# 作用:截取"产线"列每个单元格的前2个字符,生成新列"模块"
# 业务:例如"定子1线" → "定子"
dn["模块"] = dn["产线"].str[:2]

# 语法:列表拼接 + 列表推导式
# 作用:重新定义列顺序,把"模块"放在第1列
cols = ["模块", "产线", "班次"] + [col for col in dn.columns if col not in ["模块", "产线", "班次"]]

# 语法:df[列顺序列表]
# 作用:应用新的列顺序
dn = dn[cols]

# ===================== 5. 自定义模块排序(核心) =====================
# 自定义模块顺序列表
# 业务:报表最终会严格按照这个顺序从上到下排列

module_sort_list = [
    "VK", "定子", "转子", "冲压", "焊洗", "成品", "装配", "焊涂",
    "综合", "曲轴", "汽缸", "活塞", "精加", "连杆", "阀板"
]

# 语法:pd.Categorical(列, categories=顺序列表, ordered=True)
# 作用:
# 1. 把普通文本列 → 分类列
# 2. categories=module_sort_list → 排序规则按这个列表
# 3. ordered=True → 启用顺序排序
# 业务:强制模块按指定顺序排列
dn["模块"] = pd.Categorical(dn["模块"], categories=module_sort_list, ordered=True)

# 语法:df["列名"].map(字典)
# 作用:把"A班"→0,"B班"→1,用于排序时让A班在B班上面
dn["班次排序"] = dn["班次"].map({"白班": 0, "夜班": 1})


# 3. ✅ 产线自然数字排序（修复空值报错）
def extract_line_number(line_name):
    """从产线名提取数字,没有数字则返回999(排在最后)"""
    # 处理空值
    if pd.isna(line_name):
        return 999
    # 转成字符串再处理
    line_name = str(line_name)
    num_str = ''.join([c for c in line_name if c.isdigit()])
    return int(num_str) if num_str else 999


# 生成排序用的数字列
dn["产线编号"] = dn["产线"].apply(extract_line_number)

# 4. 最终排序：模块 → 班次 → 产线编号（自然顺序）
dn = dn.sort_values(by=["模块", "班次排序", "产线编号"])

# 5. 删除临时排序列
dn = dn.drop(columns=["班次排序", "产线编号"])

# ===================== 6. 保存临时Excel(排序完成) =====================
temp_file = "点检统计_临时.xlsx"
# 语法:df.to_excel(文件, index=False)
# index=False → 不导出pandas自带的行号
dn.to_excel(temp_file, index=False)

# ===================== 7. 合并单元格(最后一步) =====================
# 语法:load_workbook(文件)
# 作用:打开临时Excel文件
wb = load_workbook(temp_file)
# 语法:wb.active
# 作用:获取当前活动工作表(默认第一个sheet)
ws = wb.active

# 语法:Alignment(horizontal, vertical)
# 作用:定义对齐样式:水平居中 + 垂直居中
center_align = Alignment(horizontal="center", vertical="center")

# ===================== 8. 线长列合并逻辑修改：装配模块不合并 =====================
# 列定义

module_col = 1  # 模块列(A列)
banci_col = 3  # 班次列(C列)
line_length_col = 16  # 线长列(O列)，原变量名module_col，重命名更清晰

# 合并相关变量
current_banci = None  # 当前班次，原变量名current_module，重命名避免混淆
start_row = 2  # 合并起始行，从第2行(数据行)开始

# 颜色填充定义
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 白班-黄色
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 夜班-绿色

# 循环遍历所有数据行
for row in range(2, ws.max_row + 1):
    # 获取当前行的班次、模块值
    banci_value = ws.cell(row=row, column=banci_col).value
    module_value = ws.cell(row=row, column=module_col).value

    # 1. 班次颜色填充（原逻辑保留）
    if banci_value == "白班":
        ws[f"B{row}"].fill = yellow_fill
        ws[f"C{row}"].fill = yellow_fill
    if banci_value == "夜班":
        ws[f"B{row}"].fill = green_fill
        ws[f"C{row}"].fill = green_fill

    # 2. 核心修改：装配模块的行，线长列不合并
    if module_value in ("装配"):
        # 如果之前有正在合并的非装配区间，先完成合并
        if current_banci is not None:
            ws.merge_cells(
                start_row=start_row,
                end_row=row - 1,
                start_column=line_length_col,
                end_column=line_length_col
            )
            ws.cell(start_row, line_length_col).alignment = center_align
        # 重置合并状态，装配行不参与合并
        current_banci = None
        start_row = row + 1
        # 跳过后续合并逻辑
        continue

    # 3. 非装配模块，按原逻辑按班次合并线长列
    if banci_value != current_banci:
        # 上一个区间有数据，完成合并
        if current_banci is not None:
            ws.merge_cells(
                start_row=start_row,
                end_row=row - 1,
                start_column=line_length_col,
                end_column=line_length_col
            )
            ws.cell(start_row, line_length_col).alignment = center_align
        # 更新当前班次和合并起始行
        current_banci = banci_value
        start_row = row

# 循环结束后，处理最后一个非装配的合并区间
if current_banci is not None:
    ws.merge_cells(
        start_row=start_row,
        end_row=ws.max_row,
        start_column=line_length_col,
        end_column=line_length_col
    )
    ws.cell(start_row, line_length_col).alignment = center_align

# ===================== 9. 模块列合并（原逻辑保留） =====================
# 模块列在Excel中是第1列(A列)
module_col = 1
# 记录当前正在处理的模块名称
current_module = None
# 合并开始行:从第2行开始(第1行是表头)
start_row = 2

# 语法:for row in range(开始行, 结束行+1)
# 作用:循环遍历Excel所有数据行
for row in range(2, ws.max_row + 1):
    # 语法:ws.cell(row=行号, column=列号).value
    # 作用:获取指定单元格的值
    module_value = ws.cell(row=row, column=module_col).value

    # 语法:if 值 != 变量
    # 作用:判断当前单元格模块,与上一个模块是否不同
    if module_value != current_module:
        # 如果不是第一次循环,就合并上一组
        if current_module is not None:
            # 语法:ws.merge_cells(start_row=开始行, end_row=结束行, ...)
            # 作用:合并单元格
            ws.merge_cells(start_row=start_row, end_row=row - 1, start_column=module_col, end_column=module_col)
            # 给合并后的单元格设置居中样式
            ws.cell(start_row, module_col).alignment = center_align

        # 更新当前模块为最新模块
        current_module = module_value
        # 更新新模块的起始行
        start_row = row

# 循环结束后,合并最后一组模块
ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=module_col, end_column=module_col)
# 最后一组也设置居中
ws.cell(start_row, module_col).alignment = center_align

# ===================== 10. 保存最终文件 =====================
final_file = "点检统计_最终版.xlsx"
# 保存文件
wb.save(final_file)
# 关闭文件,释放系统资源
wb.close()

print("✅ 全部执行完成！文件已保存为：", final_file)